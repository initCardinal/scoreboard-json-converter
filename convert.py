"""
Scoreboard Excel to JSON Converter
Reads Scoreboard Test.xlsx and writes output.json.

Two workbook passes:
  - formula workbook: preserves formula strings
  - data_only workbook: preserves cached values

The output has two layers:
  - raw workbook layer: every cell, column, and merged range preserved faithfully
  - scoreboard layer: normalized metrics and records, queryable by any consumer
"""

from __future__ import annotations

import sys
import argparse
import json
import re
import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string

CONVERTER_VERSION = "1.0.0"

# Row assignments in the SCOREBOARD sheet
ROW_SECTION = 1
ROW_METRIC = 2
ROW_FOCUS = 3
ROW_SOURCE = 4
ROW_ROLE = 5
ROW_TARGET_LABEL = 6
ROW_TARGET_VALUE = 7
DATE_COLUMN_INDEX = 1  # Column A holds dates

# Rows that carry dated performance data
DATA_ROW_START = 8


def main():
    args = parse_args()
    input_path = Path(args.input)
    output_path = Path(args.output)

    if not input_path.exists():
        print(f"Error: input file not found: {input_path}", file=sys.stderr)
        sys.exit(1)

    try:
        wb_formula, wb_data = load_workbooks(input_path)
    except Exception as exc:
        print(f"Error: could not load workbook: {exc}", file=sys.stderr)
        sys.exit(1)

    if not wb_formula.sheetnames:
        print("Error: workbook contains no sheets.", file=sys.stderr)
        sys.exit(1)

    result = convert_workbook(wb_formula, wb_data, input_path)

    try:
        write_json(result, output_path)
    except Exception as exc:
        print(f"Error: could not write output: {exc}", file=sys.stderr)
        sys.exit(1)

    _print_summary(result, output_path)


def parse_args():
    parser = argparse.ArgumentParser(
        description="Convert Scoreboard Test.xlsx to output.json"
    )
    parser.add_argument(
        "--input",
        default="Scoreboard Test.xlsx",
        help="Path to the source Excel file (default: 'Scoreboard Test.xlsx')",
    )
    parser.add_argument(
        "--output",
        default="output.json",
        help="Path for the output JSON file (default: output.json)",
    )
    return parser.parse_args()


def load_workbooks(path: Path):
    """Load the workbook twice: once with formulas, once with cached values."""
    wb_formula = load_workbook(str(path), data_only=False)
    wb_data = load_workbook(str(path), data_only=True)
    return wb_formula, wb_data


def convert_workbook(wb_formula, wb_data, input_path: Path) -> dict:
    warnings = []
    sheets_out = []

    for sheet_name in wb_formula.sheetnames:
        ws_formula = wb_formula[sheet_name]
        ws_data = wb_data[sheet_name]
        sheet_out = convert_sheet(ws_formula, ws_data, warnings)
        sheets_out.append(sheet_out)

    return {
        "meta": {
            "source_file": input_path.name,
            "generated_at": datetime.datetime.now(datetime.timezone.utc)
                .isoformat()
                .replace("+00:00", "Z"),
            "converter_version": CONVERTER_VERSION,
            "warnings": warnings,
        },
        "workbook": {
            "sheet_count": len(wb_formula.sheetnames),
            "sheets": wb_formula.sheetnames,
        },
        "sheets": sheets_out,
    }


def convert_sheet(ws_formula, ws_data, warnings: list) -> dict:
    name = ws_formula.title

    min_row = ws_formula.min_row or 1
    max_row = ws_formula.max_row or 1
    min_col = ws_formula.min_column or 1
    max_col = ws_formula.max_column or 1

    min_letter = get_column_letter(min_col)
    max_letter = get_column_letter(max_col)

    merged_ranges = extract_merged_ranges(ws_formula)
    columns = extract_columns(ws_formula, min_col, max_col)
    raw_cells = extract_raw_cells(ws_formula, ws_data, min_row, max_row, min_col, max_col, merged_ranges)
    spacer_cols = detect_spacer_columns(raw_cells, min_row, max_row, min_col, max_col)

    for col in columns:
        col["is_spacer"] = col["index"] in spacer_cols

    scoreboard = _build_scoreboard_layer(
        ws_formula, ws_data, raw_cells, columns, merged_ranges,
        min_row, max_row, min_col, max_col, warnings, name
    )

    return {
        "name": name,
        "dimensions": {
            "min_row": min_row,
            "max_row": max_row,
            "min_column": min_col,
            "max_column": max_col,
            "range": f"{min_letter}{min_row}:{max_letter}{max_row}",
        },
        "merged_ranges": merged_ranges,
        "columns": columns,
        "raw_cells": raw_cells,
        "scoreboard": scoreboard,
    }


def extract_columns(ws, min_col: int, max_col: int) -> list:
    """Build the columns array with hidden status and width for every column."""
    col_dims = ws.column_dimensions
    cols = []
    for idx in range(min_col, max_col + 1):
        letter = get_column_letter(idx)
        dim = col_dims.get(letter)
        hidden = bool(dim and dim.hidden)
        width = float(dim.width) if (dim and dim.width is not None) else None
        cols.append({
            "index": idx,
            "letter": letter,
            "hidden": hidden,
            "width": width,
            "is_spacer": False,  # filled in after spacer detection
        })
    return cols


def extract_merged_ranges(ws) -> list:
    """Extract merged cell ranges with their display value."""
    out = []
    for merged in ws.merged_cells.ranges:
        ref = str(merged)
        start_cell = ws.cell(row=merged.min_row, column=merged.min_col)
        value = serialize_cell_value(start_cell.value)
        out.append({
            "range": ref,
            "value": value,
            "start_cell": f"{get_column_letter(merged.min_col)}{merged.min_row}",
            "end_cell": f"{get_column_letter(merged.max_col)}{merged.max_row}",
        })
    return out


def extract_raw_cells(ws_formula, ws_data, min_row, max_row, min_col, max_col, merged_ranges) -> dict:
    """
    Build a flat dict keyed by cell coordinate.
    Every cell in the used range is included, even if blank.
    Formulas come from ws_formula; cached values from ws_data.
    """
    # Build a lookup of which cells are merge children (not the top-left anchor)
    merged_child_map = {}  # coord -> parent coord
    for mr in merged_ranges:
        start = mr["start_cell"]
        end = mr["end_cell"]
        start_col = column_index_from_string(re.sub(r"\d", "", start))
        start_row = int(re.sub(r"[A-Z]", "", start))
        end_col = column_index_from_string(re.sub(r"\d", "", end))
        end_row = int(re.sub(r"[A-Z]", "", end))
        for r in range(start_row, end_row + 1):
            for c in range(start_col, end_col + 1):
                coord = f"{get_column_letter(c)}{r}"
                is_parent = (r == start_row and c == start_col)
                merged_child_map[coord] = {
                    "is_merged": True,
                    "merged_parent": None if is_parent else start,
                }

    cells = {}
    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            letter = get_column_letter(col)
            coord = f"{letter}{row}"

            fc = ws_formula.cell(row=row, column=col)
            dc = ws_data.cell(row=row, column=col)

            raw_value = fc.value
            formula = None
            cached_value = None

            if isinstance(raw_value, str) and raw_value.startswith("="):
                formula = raw_value
                cached_value = serialize_cell_value(dc.value)
                display_value = cached_value
            else:
                display_value = serialize_cell_value(raw_value)
                cached_value = serialize_cell_value(dc.value)

            merge_info = merged_child_map.get(coord, {"is_merged": False, "merged_parent": None})

            cells[coord] = {
                "coordinate": coord,
                "row": row,
                "column": col,
                "column_letter": letter,
                "value": display_value,
                "formula": formula,
                "cached_value": cached_value,
                "data_type": fc.data_type if fc.data_type is not None else "n",
                "number_format": fc.number_format if fc.number_format else "General",
                "is_merged": merge_info["is_merged"],
                "merged_parent": merge_info["merged_parent"],
            }

    return cells


def detect_spacer_columns(raw_cells: dict, min_row, max_row, min_col, max_col) -> set:
    """
    A column is a spacer if every cell in it (across all rows) has a null value
    and no formula. Column A (date column) is excluded from spacer classification.
    """
    spacers = set()
    for col in range(min_col, max_col + 1):
        if col == DATE_COLUMN_INDEX:
            continue
        all_blank = True
        for row in range(min_row, max_row + 1):
            letter = get_column_letter(col)
            coord = f"{letter}{row}"
            cell = raw_cells.get(coord, {})
            if cell.get("value") is not None or cell.get("formula") is not None:
                all_blank = False
                break
        if all_blank:
            spacers.add(col)
    return spacers


def _build_scoreboard_layer(ws_formula, ws_data, raw_cells, columns, merged_ranges,
                             min_row, max_row, min_col, max_col, warnings, sheet_name) -> dict:
    """Assemble the normalized scoreboard layer."""
    if max_row < DATA_ROW_START:
        warnings.append(
            f"Sheet '{sheet_name}': no data rows found (expected rows >= {DATA_ROW_START})."
        )

    spacer_set = {c["index"] for c in columns if c["is_spacer"]}

    metrics = extract_scoreboard_metrics(
        raw_cells, columns, merged_ranges, min_col, max_col, spacer_set
    )
    metric_id_map = {m["id"]: m for m in metrics}

    data_rows = list(range(DATA_ROW_START, max_row + 1))
    records = extract_scoreboard_records(raw_cells, metrics, data_rows)
    formula_audit = extract_formula_audit(raw_cells, metric_id_map, data_rows)

    return {
        "header_rows": {
            "section": ROW_SECTION,
            "metric": ROW_METRIC,
            "focus": ROW_FOCUS,
            "source": ROW_SOURCE,
            "role": ROW_ROLE,
            "target_label": ROW_TARGET_LABEL,
            "target_value": ROW_TARGET_VALUE,
        },
        "date_rows": data_rows,
        "metrics": metrics,
        "records": records,
        "formula_audit": formula_audit,
    }


def _section_map_from_merged(merged_ranges, min_col, max_col) -> dict:
    """
    Build a column-index -> section-label map from row-1 merged ranges.
    Columns not covered by a merge get the raw row-1 cell value (often null).
    """
    section_by_col = {}
    for mr in merged_ranges:
        start = mr["start_cell"]
        end = mr["end_cell"]
        # Only care about row-1 merges
        if not start.endswith("1"):
            continue
        start_col = column_index_from_string(re.sub(r"\d", "", start))
        end_col = column_index_from_string(re.sub(r"\d", "", end))
        label = mr["value"]
        for c in range(start_col, end_col + 1):
            section_by_col[c] = label
    return section_by_col


def extract_scoreboard_metrics(raw_cells, columns, merged_ranges, min_col, max_col, spacer_set) -> list:
    """
    Build a metric definition for every non-spacer, non-date column that has a heading.
    Handles duplicate headings with a stable column-letter-prefixed ID.
    """
    section_by_col = _section_map_from_merged(merged_ranges, min_col, max_col)

    # Fallback: raw row-1 value for columns not covered by a merge
    for col_idx in range(min_col, max_col + 1):
        if col_idx not in section_by_col:
            letter = get_column_letter(col_idx)
            val = raw_cells.get(f"{letter}{ROW_SECTION}", {}).get("value")
            section_by_col[col_idx] = val

    metrics = []
    seen_ids = {}  # id -> count, for deduplication safety

    for col in columns:
        idx = col["index"]
        if idx == DATE_COLUMN_INDEX or idx in spacer_set:
            continue

        letter = col["letter"]
        heading_cell = raw_cells.get(f"{letter}{ROW_METRIC}", {})
        heading = heading_cell.get("value")

        if heading is None:
            continue

        # Clean up newlines in heading for display
        heading_clean = str(heading).replace("\n", " ").replace("\r", "").strip()

        metric_id = f"{letter.lower()}_{slugify(heading_clean)}"

        # Ensure uniqueness if duplicate headings exist in different columns
        if metric_id in seen_ids:
            seen_ids[metric_id] += 1
            metric_id = f"{metric_id}_{seen_ids[metric_id]}"
        else:
            seen_ids[metric_id] = 1

        focus_raw = raw_cells.get(f"{letter}{ROW_FOCUS}", {}).get("value")
        source_raw = raw_cells.get(f"{letter}{ROW_SOURCE}", {}).get("value")
        role_raw = raw_cells.get(f"{letter}{ROW_ROLE}", {}).get("value")
        target_label_raw = raw_cells.get(f"{letter}{ROW_TARGET_LABEL}", {}).get("value")
        target_value_raw = raw_cells.get(f"{letter}{ROW_TARGET_VALUE}", {}).get("value")
        number_format = raw_cells.get(f"{letter}{ROW_METRIC}", {}).get("number_format", "General")

        # Prefer the data row number format when the heading row has General
        if number_format == "General":
            for data_row in range(DATA_ROW_START, DATA_ROW_START + 3):
                nf = raw_cells.get(f"{letter}{data_row}", {}).get("number_format", "General")
                if nf and nf != "General":
                    number_format = nf
                    break

        metrics.append({
            "id": metric_id,
            "column": letter,
            "column_index": idx,
            "heading": heading_clean,
            "section": section_by_col.get(idx),
            "focus": _clean_str(focus_raw),
            "source": _clean_str(source_raw),
            "role": _clean_str(role_raw),
            "target": {
                "label": _clean_str(target_label_raw),
                "value": target_value_raw,
            },
            "hidden": col["hidden"],
            "number_format": number_format,
        })

    return metrics


def extract_scoreboard_records(raw_cells, metrics, data_rows) -> list:
    """Build one record per data row, with a value entry for every metric."""
    records = []

    for row in data_rows:
        date_cell = raw_cells.get(f"A{row}", {})
        date_value = date_cell.get("value")
        date_str = _format_date(date_value)

        values = {}
        for metric in metrics:
            letter = metric["column"]
            coord = f"{letter}{row}"
            cell = raw_cells.get(coord, {})
            cv = cell.get("cached_value")
            has_error = isinstance(cv, str) and cv.startswith("#")

            values[metric["id"]] = {
                "cell": coord,
                "value": cell.get("value"),
                "formula": cell.get("formula"),
                "cached_value": cv,
                "number_format": cell.get("number_format", "General"),
                "has_error": has_error,
            }

        records.append({
            "row": row,
            "date": date_str,
            "values": values,
        })

    return records


def extract_formula_audit(raw_cells, metric_id_map, data_rows) -> list:
    """Collect every cell with a formula across all data rows."""
    audit = []
    metric_by_col = {m["column"]: m["id"] for m in metric_id_map.values()}

    for row in data_rows:
        for cell in raw_cells.values():
            if cell["row"] != row:
                continue
            if not cell.get("formula"):
                continue
            cv = cell.get("cached_value")
            has_error = isinstance(cv, str) and cv.startswith("#")
            col_letter = cell["column_letter"]
            audit.append({
                "cell": cell["coordinate"],
                "metric_id": metric_by_col.get(col_letter),
                "formula": cell["formula"],
                "cached_value": cv,
                "has_error": has_error,
            })

    # Sort by row then column for consistent output
    audit.sort(key=lambda x: (
        int(re.sub(r"[A-Z]", "", x["cell"])),
        column_index_from_string(re.sub(r"\d", "", x["cell"]))
    ))

    return audit


def write_json(data: dict, path: Path):
    """Write the result dict to JSON with 2-space indent."""
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False, default=_json_fallback)


def slugify(text: str) -> str:
    """Convert a heading to a lowercase slug suitable for a JSON key."""
    text = text.lower().strip()
    text = re.sub(r"[^a-z0-9]+", "_", text)
    text = re.sub(r"_+", "_", text)
    return text.strip("_")


def serialize_cell_value(value):
    """
    Convert an openpyxl cell value to a JSON-safe type.
    Dates become ISO date strings. Errors are preserved as strings.
    None stays None.
    """
    if value is None:
        return None
    if isinstance(value, datetime.datetime):
        return value.date().isoformat()
    if isinstance(value, datetime.date):
        return value.isoformat()
    if isinstance(value, datetime.timedelta):
        return str(value)
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return value
    return str(value)


def _format_date(value) -> str | None:
    """Return an ISO date string from a date, datetime, or Excel serial number."""
    if value is None:
        return None
    if isinstance(value, (datetime.date, datetime.datetime)):
        if isinstance(value, datetime.datetime):
            return value.date().isoformat()
        return value.isoformat()
    # Excel serial number (days since 1899-12-30)
    if isinstance(value, (int, float)):
        try:
            delta = datetime.timedelta(days=int(value))
            base = datetime.date(1899, 12, 30)
            return (base + delta).isoformat()
        except Exception:
            return str(value)
    return str(value)


def _clean_str(value) -> str | None:
    """Strip whitespace and newlines from a string value, return None if empty."""
    if value is None:
        return None
    s = str(value).replace("\n", " ").replace("\r", "").strip()
    return s if s else None


def _json_fallback(obj):
    """Fallback serializer for types json.dump doesn't handle natively."""
    if isinstance(obj, (datetime.date, datetime.datetime)):
        return obj.isoformat()
    raise TypeError(f"Object of type {type(obj)} is not JSON serializable")


def _print_summary(result: dict, output_path: Path):
    sheet = result["sheets"][0] if result["sheets"] else None
    meta = result["meta"]
    print(f"\nWrote: {output_path}")
    if sheet:
        sb = sheet.get("scoreboard", {})
        metrics = sb.get("metrics", [])
        records = sb.get("records", [])
        audit = sb.get("formula_audit", [])
        hidden = [c for c in sheet["columns"] if c["hidden"]]
        spacers = [c for c in sheet["columns"] if c["is_spacer"]]
        merged = sheet.get("merged_ranges", [])
        errors = [a for a in audit if a.get("has_error")]

        print(f"  Sheet         : {sheet['name']}")
        print(f"  Dimensions    : {sheet['dimensions']['range']}")
        print(f"  Total columns : {sheet['dimensions']['max_column']}")
        print(f"  Metrics       : {len(metrics)}")
        print(f"  Date rows     : {len(records)}")
        print(f"  Formulas      : {len(audit)}")
        print(f"  Formula errors: {len(errors)}")
        print(f"  Hidden columns: {len(hidden)}")
        print(f"  Spacer columns: {len(spacers)}")
        print(f"  Merged ranges : {len(merged)}")
    if meta.get("warnings"):
        print("\nWarnings:")
        for w in meta["warnings"]:
            print(f"  ! {w}")


if __name__ == "__main__":
    main()
